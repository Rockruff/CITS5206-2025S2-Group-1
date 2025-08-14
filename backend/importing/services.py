import csv, io
from openpyxl import load_workbook
from django.db import transaction
from django.utils import timezone
from .models import ImportBatch, ImportRow, RowStatus, BatchStatus


def _xlsx_to_rows(fh):
    wb = load_workbook(fh, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [
        (i, {headers[j]: row[j] for j in range(len(headers))})
        for i, row in enumerate(rows[1:], start=2)
    ]


def _csv_to_rows(fh):
    content = fh.read()
    text = (
        content.decode("utf-8", errors="ignore")
        if isinstance(content, bytes)
        else content
    )
    fh.seek(0)
    return [
        (i, rec) for i, rec in enumerate(csv.DictReader(io.StringIO(text)), start=2)
    ]


def stage_rows_from_file(batch: ImportBatch):
    if not batch.uploaded_file:
        batch.status = BatchStatus.FAILED
        batch.result_message = "No file attached."
        batch.save(update_fields=["status", "result_message"])
        return
    name = (batch.uploaded_file.name or "").lower()
    batch.status = BatchStatus.PROCESSING
    batch.save(update_fields=["status"])
    try:
        with batch.uploaded_file.open("rb") as fh:
            rows = _xlsx_to_rows(fh) if name.endswith(".xlsx") else _csv_to_rows(fh)
        objs = [
            ImportRow(batch=batch, row_number=n, raw_json=raw, status=RowStatus.PENDING)
            for (n, raw) in rows
        ]
        with transaction.atomic():
            ImportRow.objects.filter(batch=batch).delete()
            ImportRow.objects.bulk_create(objs, batch_size=1000)
            batch.total_rows = len(objs)
            batch.accepted_rows = batch.error_rows = batch.skipped_rows = 0
            batch.result_message = "Staged; ready to process."
            batch.status = BatchStatus.PARTIAL if objs else BatchStatus.COMPLETED
            batch.save()
    except Exception as e:
        batch.status = BatchStatus.FAILED
        batch.result_message = f"Parse error: {e}"
        batch.save(update_fields=["status", "result_message"])


def process_pending_rows(batch: ImportBatch):
    pending = ImportRow.objects.select_for_update(skip_locked=True).filter(
        batch=batch, status=RowStatus.PENDING
    )
    with transaction.atomic():
        for row in pending:
            data = row.raw_json or {}
            has_email = bool(str(data.get("email") or data.get("Email") or "").strip())
            has_course = bool(
                str(data.get("course") or data.get("Course") or "").strip()
            )
            if has_email and has_course:
                row.status = RowStatus.ACCEPTED
                row.action_taken = "validated"
            else:
                row.status = RowStatus.ERROR
                row.error_details = "Missing required fields: email and/or course."
            row.processed_at = timezone.now()
            row.save()
        accepted = ImportRow.objects.filter(
            batch=batch, status=RowStatus.ACCEPTED
        ).count()
        errors = ImportRow.objects.filter(batch=batch, status=RowStatus.ERROR).count()
        skipped = ImportRow.objects.filter(
            batch=batch, status=RowStatus.SKIPPED
        ).count()
        batch.accepted_rows, batch.error_rows, batch.skipped_rows = (
            accepted,
            errors,
            skipped,
        )
        batch.status = (
            BatchStatus.COMPLETED
            if (batch.total_rows == accepted + errors + skipped)
            else BatchStatus.PARTIAL
        )
        batch.processed_at = timezone.now()
        batch.result_message = "Processed pending rows."
        batch.save()
