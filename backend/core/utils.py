import csv
from typing import List, Tuple, IO, Sequence
from openpyxl import load_workbook

from django.utils import timezone
from django.utils.dateparse import parse_datetime, parse_date
from datetime import datetime


# Common column names
NAME_COL = "Name"
UID_COL = "UserID"

COMPLETEION_DATE_COL = "Completion Date"
SCORE_COL = "Score"


def parse_xlsx(file: IO[bytes], columns: Sequence[str]) -> List[Tuple[int, ...]]:
    """
    Parse an .xlsx file and return rows as (row_index, ...columns...).
    Example: parse_xlsx(file, [UID_COL, NAME_COL])
    """
    ws = load_workbook(file, read_only=True, data_only=True).active
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    idxs = [headers.index(col) for col in columns]

    return [
        (row_idx, *(str(row[i]).strip() for i in idxs))
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2)
    ]


def parse_csv(file: IO[bytes], columns: Sequence[str]) -> List[Tuple[int, ...]]:
    """
    Parse a .csv file and return rows as (row_index, ...columns...).
    Example: parse_csv(file, [UID_COL, NAME_COL])
    """
    reader = csv.DictReader(file.read().decode("utf-8").splitlines())
    return [
        (row_idx, *(row[col].strip() for col in columns))
        for row_idx, row in enumerate(reader, start=2)  # start=2 for consistency with Excel
    ]


def paginate_qs(qs, query_params, page_size_default, Serializer, Response):
    try:
        page = int(query_params.get("page"))
        assert page >= 1
    except Exception:
        page = 1
    try:
        page_size = int(query_params.get("page_size"))
        assert page_size >= 1
    except Exception:
        page_size = page_size_default

    total_items = len(qs) if isinstance(qs, list) else qs.count()
    total_pages = (total_items + page_size - 1) // page_size
    start = (page - 1) * page_size
    end = start + page_size
    items = [Serializer(v).data for v in qs[start:end]]

    return Response(
        {
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
            "total_items": total_items,
            "items": items,
        }
    )


def parse_to_aware_datetime(value):
    """
    Safely parse a string into a timezone-aware datetime.

    - Accepts full datetime or date strings.
    - If timezone info is missing, assumes the current Django timezone.
    - Returns None if input is invalid or empty.
    """
    if not value:
        return None

    try:
        # Try to parse as full datetime string first
        dt = parse_datetime(value)
        if not dt:
            # Fallback: try to parse as a date string
            d = parse_date(value)
            if not d:
                return None
            dt = datetime.combine(d, datetime.min.time())

        # Make timezone-aware if it's naive
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_default_timezone())

        return dt
    except Exception:
        # Catch any unexpected parsing or type errors
        return None
