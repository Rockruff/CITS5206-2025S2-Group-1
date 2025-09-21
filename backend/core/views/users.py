import csv
import io
import re
from typing import List, Dict

from django.db import transaction
from django.http import Http404
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

from core.models import User, UserAlias
from core.models import UserGroup
from core.serializers.users import (
    UserSerializer,
    UserCreateSerializer,
    UserUpdateSerializer,
    UserAliasCreateSerializer,
    UserAliasDeleteSerializer,
)
from core.permissions import IsAuthenticated, IsAdmin

try:
    import openpyxl

    HAS_XLSX = True
except Exception:
    HAS_XLSX = False


class UserViewSet(viewsets.GenericViewSet):
    permission_classes = [IsAdmin]

    def get_object(self):
        pk = self.kwargs.get("pk")
        try:
            alias = UserAlias.objects.get(id=pk)
        except UserAlias.DoesNotExist:
            raise Http404
        return alias.user

    # POST /users
    def create(self, request):
        serializer = UserCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response(UserSerializer(user).data)

    # GET /users/{id}
    def retrieve(self, request, *args, **kwargs):
        user = self.get_object()
        return Response(UserSerializer(user).data)

    # GET /users
    def list(self, request):
        qs = User.objects.all()

        # Filters
        role = request.query_params.get("role")
        if role:
            qs = qs.filter(role=role)

        id_kw = request.query_params.get("id")
        if id_kw:
            qs = qs.filter(aliases__id__icontains=id_kw).distinct()

        name_kw = request.query_params.get("name")
        if name_kw:
            keywords = name_kw.split()
            for kw in keywords:
                qs = qs.filter(name__icontains=kw)

        group_id = request.query_params.get("group")
        if group_id:
            qs = qs.filter(groups__id=group_id)

        # Ordering
        order_by = request.query_params.get("order_by", "id")
        if order_by not in {"id", "-id", "name", "-name", "role", "-role"}:
            return Response({"error": "Invalid order_by field"}, status=400)
        qs = qs.order_by(order_by)

        # Pagination
        try:
            page = int(request.query_params.get("page", "1"))
            page_size = int(request.query_params.get("page_size", "10"))
            assert page >= 1 and page_size >= 1
        except Exception:
            return Response({"error": "Invalid pagination parameters"}, status=400)

        total_items = qs.count()
        total_pages = (total_items + page_size - 1) // page_size
        start = (page - 1) * page_size
        end = start + page_size
        items = [UserSerializer(u).data for u in qs[start:end]]

        return Response(
            {
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
                "total_items": total_items,
                "items": items,
            }
        )

    # PATCH /users/{id}
    def partial_update(self, request, *args, **kwargs):
        user = self.get_object()
        serializer = UserUpdateSerializer(instance=user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UserSerializer(user).data)

    # DELETE /users/{id}
    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        user.delete()
        return Response()

    # GET /users/me
    @action(detail=False, methods=["get"], url_path="me", permission_classes=[IsAuthenticated])
    def me(self, request):
        return Response(UserSerializer(request.user).data)

    @action(detail=True, methods=["put"], url_path="groups")
    def set_groups(self, request, *args, **kwargs):
        """
        Replace this user's group memberships with the given list of group UUIDs.
        Body: { "groups": ["<uuid>", ...] }
        """
        user = self.get_object()
        group_ids = request.data.get("groups", [])
        if not isinstance(group_ids, list):
            return Response({"error": "groups must be a list of UUIDs"}, status=400)

        # Validate IDs exist; you can soften this if you prefer
        groups = list(UserGroup.objects.filter(id__in=group_ids))
        if len(groups) != len(set(group_ids)):
            return Response({"error": "one or more groups not found"}, status=400)

        # This works because the M2M is defined on UserGroup with related_name="groups"
        user.groups.set(groups)  # replace all memberships at once
        return Response(UserSerializer(user).data, status=status.HTTP_200_OK)

    # POST /users/{id}/aliases
    # DELETE /users/{id}/aliases
    # Since I want to use the same endpoint for both adding and removing aliases,
    # I will check the request method to determine the action.
    @action(detail=True, methods=["post", "delete"], url_path="aliases")
    def add_alias(self, request, *args, **kwargs):
        user = self.get_object()
        serializer_class = (
            UserAliasCreateSerializer if request.method == "POST" else UserAliasDeleteSerializer
        )
        serializer = serializer_class(instance=user, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(UserSerializer(user).data)

    # POST /users/batch
    @action(detail=False, methods=["post"], url_path="batch")
    def batch(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "Missing file"}, status=400)

        # Determine file type
        name = file.name.lower()
        try:
            rows = []
            if name.endswith(".csv"):
                text = file.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(text))
                for r in reader:
                    rows.append(r)
            elif name.endswith(".xlsx"):
                if not HAS_XLSX:
                    return Response({"error": "XLSX support not available on server"}, status=400)
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                headers = [
                    str(c.value).strip() if c.value is not None else ""
                    for c in next(ws.iter_rows(min_row=1, max_row=1))
                ]

                def normalise(h):
                    return re.sub(r"\W+", "", h).lower()

                header_map = {normalise(h): idx for idx, h in enumerate(headers)}
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    rec = {}
                    for norm, idx in header_map.items():
                        rec[headers[idx]] = row[idx] if idx < len(row) else None
                    rows.append(rec)
            else:
                return Response({"error": "Unsupported file type (use .csv or .xlsx)"}, status=400)
        except Exception:
            return Response({"error": "Failed to parse file"}, status=400)

        def pick(d: Dict, keys: List[str]):
            # keys tested case-insensitive and lenient
            norm = {re.sub(r"\W+", "", k or "").lower(): v for k, v in d.items()}
            id_val = None
            name_val = None
            for k in ["uwaid", "id", "uw a id", "uwahid", "uw a_id"]:
                if k in norm:
                    id_val = norm[k]
                    break
            for k in ["name", "fullname", "full_name"]:
                if k in norm:
                    name_val = norm[k]
                    break
            return (
                str(id_val).strip() if id_val is not None else None,
                str(name_val).strip() if name_val is not None else None,
            )

        errors = []
        to_create = []
        skipped = 0
        created = 0

        with transaction.atomic():
            for idx, row in enumerate(rows, start=1):
                uid, uname = pick(row, [])
                if not uid or not re.match(r"^\d{8}$", uid):
                    errors.append({"row": idx, "msg": "Invalid UWA ID"})
                    continue
                if not uname:
                    errors.append({"row": idx, "msg": "Missing Name"})
                    continue

                try:
                    existing = User.objects.get(id=uid)
                    if existing.name != uname:
                        errors.append({"row": idx, "msg": f"Name mismatch for UWA ID {uid}"})
                    else:
                        skipped += 1
                except User.DoesNotExist:
                    u = User(id=uid, name=uname)
                    to_create.append(u)

            if to_create:
                User.objects.bulk_create(to_create)
                # create aliases for new users
                alias_objs = [UserAlias(user=u, id=u.id) for u in to_create]
                UserAlias.objects.bulk_create(alias_objs)
                created = len(to_create)

        return Response({"errors": errors, "created_count": created, "skipped_count": skipped})
